"""Consent Registry — regulatory consent texts, sync, and audit.

Loads consent checkbox definitions from an audit-ready Excel workbook,
stores them in PostgreSQL, and keeps them in sync via a periodic background
task.  The admin portal reads from this registry to display current consent
requirements per product, AI rules, and implementation mapping.
"""

import asyncio
import hashlib
import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, HTTPException, Request
from sqlalchemy import text

logger = logging.getLogger(__name__)

router = APIRouter()

CONSENT_XLSX_PATH = os.getenv(
    "CONSENT_XLSX_PATH",
    str(Path(__file__).resolve().parents[3] / "Consent_Checkbox_Texts_Audit_Ready 1.xlsx"),
)
SYNC_INTERVAL_HOURS = int(os.getenv("CONSENT_SYNC_INTERVAL_HOURS", "6"))
SOURCE_CHECK_INTERVAL_HOURS = int(os.getenv("SOURCE_CHECK_INTERVAL_HOURS", "24"))


# ── DB helpers ──────────────────────────────────────────────────

async def _ensure_tables(session):
    """Create consent registry tables if they don't exist."""
    await session.execute(text("""
        CREATE TABLE IF NOT EXISTS consent_sync_log (
            id SERIAL PRIMARY KEY,
            version INTEGER NOT NULL,
            file_hash VARCHAR(64) NOT NULL,
            source_file VARCHAR(500) NOT NULL,
            sheets_synced JSONB NOT NULL DEFAULT '{}',
            status VARCHAR(20) NOT NULL DEFAULT 'success',
            error_message TEXT,
            synced_at TIMESTAMP DEFAULT NOW(),
            synced_by VARCHAR(100) DEFAULT 'system'
        )
    """))
    await session.execute(text("""
        CREATE TABLE IF NOT EXISTS consent_texts (
            id SERIAL PRIMARY KEY,
            code VARCHAR(50) NOT NULL,
            checkbox_name VARCHAR(500) NOT NULL,
            ui_text TEXT NOT NULL,
            help_text TEXT,
            location TEXT,
            mandatory_to_tick VARCHAR(200),
            when_applicable TEXT,
            notes TEXT,
            sync_version INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(code)
        )
    """))
    await session.execute(text("""
        CREATE TABLE IF NOT EXISTS product_consent_map (
            id SERIAL PRIMARY KEY,
            product_name VARCHAR(200) NOT NULL,
            required_in_ui VARCHAR(500),
            checkbox_code VARCHAR(50) NOT NULL,
            ui_text TEXT,
            mandatory_to_tick VARCHAR(200),
            location TEXT,
            when_shown TEXT,
            ai_consent_needed TEXT,
            regulatory_note TEXT,
            implementation_note TEXT,
            sync_version INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """))
    await session.execute(text("""
        CREATE TABLE IF NOT EXISTS ai_consent_rules (
            id SERIAL PRIMARY KEY,
            ai_use_case VARCHAR(500) NOT NULL,
            consent_needed VARCHAR(500),
            location TEXT,
            can_ai_run_without VARCHAR(200),
            fallback TEXT,
            products TEXT,
            notes TEXT,
            sync_version INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """))
    await session.execute(text("""
        CREATE TABLE IF NOT EXISTS consent_implementation_map (
            id SERIAL PRIMARY KEY,
            checkbox_code VARCHAR(50) NOT NULL,
            db_fields TEXT,
            capture_point TEXT,
            withdrawal_location TEXT,
            enforcement_rule TEXT,
            audit_evidence TEXT,
            owner VARCHAR(200),
            notes TEXT,
            sync_version INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(checkbox_code)
        )
    """))
    await session.execute(text("""
        CREATE TABLE IF NOT EXISTS regulatory_sources (
            id SERIAL PRIMARY KEY,
            instrument VARCHAR(500) NOT NULL,
            url TEXT NOT NULL,
            notes TEXT,
            sync_version INTEGER NOT NULL,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """))
    await session.execute(text("""
        CREATE TABLE IF NOT EXISTS regulatory_source_checks (
            id SERIAL PRIMARY KEY,
            instrument VARCHAR(500) NOT NULL,
            url TEXT NOT NULL UNIQUE,
            content_hash VARCHAR(64),
            previous_hash VARCHAR(64),
            status VARCHAR(20) NOT NULL DEFAULT 'pending',
            http_status INTEGER,
            content_length INTEGER,
            last_changed_at TIMESTAMP,
            error_message TEXT,
            checked_at TIMESTAMP DEFAULT NOW(),
            reviewed_at TIMESTAMP,
            reviewed_by VARCHAR(100)
        )
    """))
    await session.commit()


# ── Sync logic ──────────────────────────────────────────────────

def _file_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _read_xlsx(path: str) -> dict:
    """Read all sheets from the consent workbook into plain dicts."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    data = {}

    # Official_Texts: row 2 = headers, rows 3+ = data
    if "Official_Texts" in wb.sheetnames:
        ws = wb["Official_Texts"]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        data["official_texts"] = [
            {
                "code": r[0], "checkbox_name": r[1], "ui_text": r[2],
                "help_text": r[3], "location": r[4],
                "mandatory_to_tick": r[5], "when_applicable": r[6], "notes": r[7],
            }
            for r in rows if r and r[0]
        ]

    # Product_Checkboxes: row 2 = headers, rows 3+ = data
    if "Product_Checkboxes" in wb.sheetnames:
        ws = wb["Product_Checkboxes"]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        data["product_checkboxes"] = [
            {
                "product_name": r[0], "required_in_ui": r[1],
                "checkbox_code": r[2], "ui_text": r[3],
                "mandatory_to_tick": r[4], "location": r[5],
                "when_shown": r[6], "ai_consent_needed": r[7],
                "regulatory_note": r[8],
                "implementation_note": r[9] if len(r) > 9 else None,
            }
            for r in rows if r and r[0]
        ]

    # AI_Rules: row 2 = headers, rows 3+ = data
    if "AI_Rules" in wb.sheetnames:
        ws = wb["AI_Rules"]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        data["ai_rules"] = [
            {
                "ai_use_case": r[0], "consent_needed": r[1],
                "location": r[2], "can_ai_run_without": r[3],
                "fallback": r[4], "products": r[5], "notes": r[6],
            }
            for r in rows if r and r[0]
        ]

    # Implementation_Map: row 2 = headers, rows 3+ = data
    if "Implementation_Map" in wb.sheetnames:
        ws = wb["Implementation_Map"]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        data["implementation_map"] = [
            {
                "checkbox_code": r[0], "db_fields": r[1],
                "capture_point": r[2], "withdrawal_location": r[3],
                "enforcement_rule": r[4], "audit_evidence": r[5],
                "owner": r[6], "notes": r[7] if len(r) > 7 else None,
            }
            for r in rows if r and r[0]
        ]

    # Sources: row 2 = headers, rows 3+ = data
    if "Sources" in wb.sheetnames:
        ws = wb["Sources"]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        data["sources"] = [
            {"instrument": r[0], "url": r[1], "notes": r[2] if len(r) > 2 else None}
            for r in rows if r and r[0]
        ]

    wb.close()
    return data


async def _do_sync(session_factory, *, synced_by: str = "system", force: bool = False):
    """Sync the consent xlsx into the database.

    Returns a dict with sync result details.
    """
    xlsx_path = CONSENT_XLSX_PATH
    if not os.path.isfile(xlsx_path):
        return {"status": "skipped", "reason": f"File not found: {xlsx_path}"}

    file_hash = _file_sha256(xlsx_path)

    async with session_factory() as session:
        await _ensure_tables(session)

        # Check if already synced with this hash
        if not force:
            result = await session.execute(
                text("SELECT file_hash, version FROM consent_sync_log ORDER BY id DESC LIMIT 1")
            )
            last = result.mappings().fetchone()
            if last and last["file_hash"] == file_hash:
                return {
                    "status": "no_change",
                    "version": last["version"],
                    "file_hash": file_hash,
                    "message": "File unchanged since last sync",
                }

        # Read the xlsx
        data = _read_xlsx(xlsx_path)

        # Determine next version
        result = await session.execute(
            text("SELECT COALESCE(MAX(version), 0) AS max_v FROM consent_sync_log")
        )
        next_version = result.scalar() + 1

        sheets_synced = {}

        # ── Official texts ──
        texts = data.get("official_texts", [])
        if texts:
            await session.execute(text("DELETE FROM consent_texts"))
            for t in texts:
                await session.execute(
                    text("""INSERT INTO consent_texts
                        (code, checkbox_name, ui_text, help_text, location,
                         mandatory_to_tick, when_applicable, notes, sync_version)
                        VALUES (:code, :name, :ui, :help, :loc, :mand, :when, :notes, :v)"""),
                    {"code": t["code"], "name": t["checkbox_name"],
                     "ui": t["ui_text"], "help": t["help_text"],
                     "loc": t["location"], "mand": t["mandatory_to_tick"],
                     "when": t["when_applicable"], "notes": t["notes"],
                     "v": next_version},
                )
            sheets_synced["official_texts"] = len(texts)

        # ── Product consent map ──
        pcm = data.get("product_checkboxes", [])
        if pcm:
            await session.execute(text("DELETE FROM product_consent_map"))
            for p in pcm:
                await session.execute(
                    text("""INSERT INTO product_consent_map
                        (product_name, required_in_ui, checkbox_code, ui_text,
                         mandatory_to_tick, location, when_shown, ai_consent_needed,
                         regulatory_note, implementation_note, sync_version)
                        VALUES (:prod, :req, :code, :ui, :mand, :loc, :when, :ai,
                                :reg, :impl, :v)"""),
                    {"prod": p["product_name"], "req": p["required_in_ui"],
                     "code": p["checkbox_code"], "ui": p["ui_text"],
                     "mand": p["mandatory_to_tick"], "loc": p["location"],
                     "when": p["when_shown"], "ai": p["ai_consent_needed"],
                     "reg": p["regulatory_note"], "impl": p["implementation_note"],
                     "v": next_version},
                )
            sheets_synced["product_checkboxes"] = len(pcm)

        # ── AI rules ──
        ai = data.get("ai_rules", [])
        if ai:
            await session.execute(text("DELETE FROM ai_consent_rules"))
            for r in ai:
                await session.execute(
                    text("""INSERT INTO ai_consent_rules
                        (ai_use_case, consent_needed, location, can_ai_run_without,
                         fallback, products, notes, sync_version)
                        VALUES (:use, :consent, :loc, :can, :fb, :prods, :notes, :v)"""),
                    {"use": r["ai_use_case"], "consent": r["consent_needed"],
                     "loc": r["location"], "can": r["can_ai_run_without"],
                     "fb": r["fallback"], "prods": r["products"],
                     "notes": r["notes"], "v": next_version},
                )
            sheets_synced["ai_rules"] = len(ai)

        # ── Implementation map ──
        impl = data.get("implementation_map", [])
        if impl:
            await session.execute(text("DELETE FROM consent_implementation_map"))
            for m in impl:
                await session.execute(
                    text("""INSERT INTO consent_implementation_map
                        (checkbox_code, db_fields, capture_point, withdrawal_location,
                         enforcement_rule, audit_evidence, owner, notes, sync_version)
                        VALUES (:code, :db, :cap, :wd, :enf, :aud, :own, :notes, :v)"""),
                    {"code": m["checkbox_code"], "db": m["db_fields"],
                     "cap": m["capture_point"], "wd": m["withdrawal_location"],
                     "enf": m["enforcement_rule"], "aud": m["audit_evidence"],
                     "own": m["owner"], "notes": m["notes"],
                     "v": next_version},
                )
            sheets_synced["implementation_map"] = len(impl)

        # ── Regulatory sources ──
        src = data.get("sources", [])
        if src:
            await session.execute(text("DELETE FROM regulatory_sources"))
            for s in src:
                await session.execute(
                    text("""INSERT INTO regulatory_sources
                        (instrument, url, notes, sync_version)
                        VALUES (:inst, :url, :notes, :v)"""),
                    {"inst": s["instrument"], "url": s["url"],
                     "notes": s["notes"], "v": next_version},
                )
            sheets_synced["sources"] = len(src)

        # ── Log the sync ──
        total = sum(sheets_synced.values())
        await session.execute(
            text("""INSERT INTO consent_sync_log
                (version, file_hash, source_file, sheets_synced, status, synced_by)
                VALUES (:v, :hash, :file, CAST(:sheets AS jsonb), 'success', :by)"""),
            {"v": next_version, "hash": file_hash,
             "file": os.path.basename(xlsx_path),
             "sheets": json.dumps(sheets_synced), "by": synced_by},
        )
        await session.commit()

        logger.info("Consent registry synced: version=%d, records=%d, hash=%s",
                     next_version, total, file_hash[:12])
        return {
            "status": "synced",
            "version": next_version,
            "file_hash": file_hash,
            "records": total,
            "sheets": sheets_synced,
        }


async def start_background_sync(session_factory):
    """Background loop — run initial sync then repeat every SYNC_INTERVAL_HOURS."""
    # Small delay to let the API finish starting
    await asyncio.sleep(5)

    while True:
        try:
            result = await _do_sync(session_factory)
            logger.info("Background consent sync: %s", result.get("status"))
        except Exception as e:
            logger.error("Background consent sync failed: %s", e)
        await asyncio.sleep(SYNC_INTERVAL_HOURS * 3600)


# ── Regulatory source change detection ──────────────────────────

async def _check_single_source(url: str) -> dict:
    """Fetch a URL and return hash + metadata. Does NOT store anything."""
    import httpx

    try:
        async with httpx.AsyncClient(
            timeout=30.0,
            follow_redirects=True,
            headers={"User-Agent": "BankOfferAI-RegulatoryMonitor/1.0"},
        ) as client:
            resp = await client.get(url)
            body = resp.content
            content_hash = hashlib.sha256(body).hexdigest()
            return {
                "content_hash": content_hash,
                "http_status": resp.status_code,
                "content_length": len(body),
                "error": None,
            }
    except Exception as e:
        return {
            "content_hash": None,
            "http_status": None,
            "content_length": None,
            "error": str(e)[:500],
        }


async def _check_all_sources(session_factory):
    """Check all regulatory source URLs for content changes."""
    async with session_factory() as session:
        await _ensure_tables(session)

        # Get all registered sources
        result = await session.execute(
            text("SELECT instrument, url FROM regulatory_sources ORDER BY id")
        )
        sources = result.mappings().fetchall()

        if not sources:
            return {"status": "skipped", "reason": "No regulatory sources in DB"}

        results = []
        for src in sources:
            url = src["url"]
            instrument = src["instrument"]

            check = await _check_single_source(url)

            if check["error"]:
                # Fetch failed
                await session.execute(
                    text("""INSERT INTO regulatory_source_checks
                        (instrument, url, status, error_message, checked_at)
                        VALUES (:inst, :url, 'error', :err, NOW())
                        ON CONFLICT (url) DO UPDATE SET
                            status = 'error',
                            error_message = :err,
                            http_status = NULL,
                            checked_at = NOW()"""),
                    {"inst": instrument, "url": url, "err": check["error"]},
                )
                results.append({"url": url, "status": "error", "error": check["error"]})
                continue

            # Check existing record
            existing = await session.execute(
                text("SELECT content_hash, status FROM regulatory_source_checks WHERE url = :url"),
                {"url": url},
            )
            row = existing.mappings().fetchone()

            if not row:
                # First check — establish baseline
                await session.execute(
                    text("""INSERT INTO regulatory_source_checks
                        (instrument, url, content_hash, status, http_status,
                         content_length, checked_at)
                        VALUES (:inst, :url, :hash, 'ok', :http, :len, NOW())
                        ON CONFLICT (url) DO UPDATE SET
                            instrument = :inst, content_hash = :hash,
                            status = 'ok', http_status = :http,
                            content_length = :len, checked_at = NOW()"""),
                    {"inst": instrument, "url": url,
                     "hash": check["content_hash"],
                     "http": check["http_status"],
                     "len": check["content_length"]},
                )
                results.append({"url": url, "status": "initial"})
            elif check["content_hash"] != row["content_hash"]:
                # Content changed!
                await session.execute(
                    text("""UPDATE regulatory_source_checks SET
                        previous_hash = content_hash,
                        content_hash = :hash,
                        status = 'changed',
                        http_status = :http,
                        content_length = :len,
                        last_changed_at = NOW(),
                        checked_at = NOW(),
                        error_message = NULL
                        WHERE url = :url"""),
                    {"hash": check["content_hash"],
                     "http": check["http_status"],
                     "len": check["content_length"],
                     "url": url},
                )
                logger.warning("Regulatory source CHANGED: %s (%s)", instrument, url)
                results.append({"url": url, "status": "changed", "instrument": instrument})
            else:
                # No change — update checked_at, keep status (ok or already-reviewed)
                new_status = "ok" if row["status"] in ("ok", "initial") else row["status"]
                await session.execute(
                    text("""UPDATE regulatory_source_checks SET
                        http_status = :http,
                        content_length = :len,
                        checked_at = NOW(),
                        error_message = NULL,
                        status = CASE WHEN status = 'error' THEN 'ok' ELSE status END
                        WHERE url = :url"""),
                    {"http": check["http_status"],
                     "len": check["content_length"],
                     "url": url},
                )
                results.append({"url": url, "status": "unchanged"})

        await session.commit()

        changed = [r for r in results if r["status"] == "changed"]
        errors = [r for r in results if r["status"] == "error"]
        logger.info("Source check complete: %d sources, %d changed, %d errors",
                     len(results), len(changed), len(errors))
        return {
            "status": "checked",
            "total": len(results),
            "changed": len(changed),
            "errors": len(errors),
            "details": results,
        }


async def start_background_source_checks(session_factory):
    """Background loop — check regulatory source URLs periodically."""
    # Wait for initial consent sync to populate sources first
    await asyncio.sleep(30)

    while True:
        try:
            result = await _check_all_sources(session_factory)
            logger.info("Background source check: %s", result.get("status"))
        except Exception as e:
            logger.error("Background source check failed: %s", e)
        await asyncio.sleep(SOURCE_CHECK_INTERVAL_HOURS * 3600)


# ── API endpoints ───────────────────────────────────────────────

@router.get("/sync-status", summary="Get consent registry sync status")
async def get_sync_status(request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)

        result = await session.execute(
            text("""SELECT id, version, file_hash, source_file, sheets_synced,
                           status, error_message, synced_at, synced_by
                    FROM consent_sync_log ORDER BY id DESC LIMIT 10""")
        )
        rows = result.mappings().fetchall()

        if not rows:
            return {
                "current": None,
                "history": [],
                "next_sync_hours": SYNC_INTERVAL_HOURS,
                "source_file": os.path.basename(CONSENT_XLSX_PATH),
            }

        current = dict(rows[0])
        current["synced_at"] = current["synced_at"].isoformat() if current["synced_at"] else None

        history = []
        for r in rows:
            entry = dict(r)
            entry["synced_at"] = entry["synced_at"].isoformat() if entry["synced_at"] else None
            history.append(entry)

        return {
            "current": current,
            "history": history,
            "next_sync_hours": SYNC_INTERVAL_HOURS,
            "source_file": os.path.basename(CONSENT_XLSX_PATH),
        }


@router.post("/sync", summary="Trigger manual consent registry sync")
async def trigger_sync(request: Request, force: bool = False):
    session_factory = request.app.state.db_session_factory
    try:
        result = await _do_sync(session_factory, synced_by="admin", force=force)
        return result
    except Exception as e:
        logger.error("Manual consent sync failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Sync failed: {e}")


@router.get("/texts", summary="Official consent checkbox texts")
async def get_consent_texts(request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        result = await session.execute(
            text("SELECT * FROM consent_texts ORDER BY code")
        )
        return [dict(r) for r in result.mappings().fetchall()]


@router.get("/product-map", summary="Per-product consent checkbox requirements")
async def get_product_consent_map(request: Request, product: str | None = None):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        query = "SELECT * FROM product_consent_map"
        params = {}
        if product:
            query += " WHERE product_name = :prod"
            params["prod"] = product
        query += " ORDER BY product_name, checkbox_code"
        result = await session.execute(text(query), params)
        return [dict(r) for r in result.mappings().fetchall()]


@router.get("/product-map/products", summary="List distinct products in consent map")
async def get_consent_products(request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        result = await session.execute(
            text("SELECT DISTINCT product_name FROM product_consent_map ORDER BY product_name")
        )
        return [r["product_name"] for r in result.mappings().fetchall()]


@router.get("/ai-rules", summary="AI-specific consent rules")
async def get_ai_rules(request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        result = await session.execute(
            text("SELECT * FROM ai_consent_rules ORDER BY id")
        )
        return [dict(r) for r in result.mappings().fetchall()]


@router.get("/implementation-map", summary="Consent implementation mapping (DB/UI)")
async def get_implementation_map(request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        result = await session.execute(
            text("SELECT * FROM consent_implementation_map ORDER BY checkbox_code")
        )
        return [dict(r) for r in result.mappings().fetchall()]


@router.get("/sources", summary="Regulatory source references")
async def get_sources(request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        result = await session.execute(
            text("SELECT * FROM regulatory_sources ORDER BY id")
        )
        return [dict(r) for r in result.mappings().fetchall()]


@router.get("/source-checks", summary="Regulatory source change detection status")
async def get_source_checks(request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        result = await session.execute(
            text("""SELECT id, instrument, url, content_hash, previous_hash,
                           status, http_status, content_length,
                           last_changed_at, error_message, checked_at,
                           reviewed_at, reviewed_by
                    FROM regulatory_source_checks ORDER BY
                        CASE status WHEN 'changed' THEN 0 WHEN 'error' THEN 1 ELSE 2 END,
                        instrument""")
        )
        rows = result.mappings().fetchall()
        checks = []
        for r in rows:
            entry = dict(r)
            for ts_field in ("last_changed_at", "checked_at", "reviewed_at"):
                if entry[ts_field]:
                    entry[ts_field] = entry[ts_field].isoformat()
            checks.append(entry)

        changed_count = sum(1 for c in checks if c["status"] == "changed")
        error_count = sum(1 for c in checks if c["status"] == "error")
        return {
            "checks": checks,
            "total": len(checks),
            "changed": changed_count,
            "errors": error_count,
            "check_interval_hours": SOURCE_CHECK_INTERVAL_HOURS,
        }


@router.post("/check-sources", summary="Trigger manual regulatory source check")
async def trigger_source_check(request: Request):
    session_factory = request.app.state.db_session_factory
    try:
        result = await _check_all_sources(session_factory)
        return result
    except Exception as e:
        logger.error("Manual source check failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Source check failed: {e}")


@router.post(
    "/source-checks/{check_id}/review",
    summary="Mark a changed source as reviewed (acknowledge the change)",
)
async def review_source_check(check_id: int, request: Request):
    session_factory = request.app.state.db_session_factory
    async with session_factory() as session:
        await _ensure_tables(session)
        result = await session.execute(
            text("SELECT id, status FROM regulatory_source_checks WHERE id = :id"),
            {"id": check_id},
        )
        row = result.mappings().fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Check not found")

        await session.execute(
            text("""UPDATE regulatory_source_checks SET
                status = 'ok',
                reviewed_at = NOW(),
                reviewed_by = 'admin',
                previous_hash = NULL
                WHERE id = :id"""),
            {"id": check_id},
        )
        await session.commit()
        return {"status": "reviewed", "check_id": check_id}
